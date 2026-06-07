"""
Service to parse uploaded Excel files containing purchase orders.
Extracts orders and their line items, then creates ProcessedOrder records.
"""
from datetime import datetime
from typing import Any
from uuid import UUID

import openpyxl
from sqlalchemy.orm import Session

from app.models.document import OrderLine, ProcessedOrder, RawDocument
from app.services.mapping_service import resolve_temp_code


def parse_excel_orders(file_path: str) -> list[dict[str, Any]]:
    """
    Parse an Excel file and extract orders with line items.
    Supports common PO Excel formats:
    - Each sheet = one order, or
    - Single sheet with multiple orders separated by headers
    
    Returns list of order dicts with 'lines' key.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    orders = []

    try:
        from app.services.excel_order_parser import _is_satori_template, parse_excel_order

        sample_rows = list(wb.active.iter_rows(min_row=1, max_row=25, values_only=True))
        if _is_satori_template(sample_rows):
            parsed = parse_excel_order(file_path)
            lines = [
                {
                    "product_name": item.get("product_name"),
                    "product_code": item.get("product_code"),
                    "quantity": item.get("quantity"),
                    "unit_price": item.get("unit_price"),
                    "uom": item.get("uom"),
                    "line_total": item.get("line_total"),
                    "tax_rate": item.get("tax_rate"),
                    "discount_rate": item.get("discount_rate"),
                    "discount_amount": item.get("discount_amount"),
                }
                for item in parsed.get("items", [])
                if _to_number(item.get("quantity")) and _to_number(item.get("quantity")) > 0
            ]
            sheet_name = wb.active.title
            wb.close()
            return [{
                "sheet_name": sheet_name,
                "lines": lines,
                "total_amount": parsed.get("total_amount") or sum(l["line_total"] or 0 for l in lines),
            }] if lines else []
    except Exception:
        pass

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Try to detect header row (look for common column names)
        header_idx = _find_header_row(rows)
        if header_idx is None:
            continue

        headers = [str(c).strip().lower() if c else "" for c in rows[header_idx]]
        col_map = _map_columns(headers)

        if not col_map.get("product_name"):
            continue

        # Extract lines from rows after header
        lines = []
        for row in rows[header_idx + 1:]:
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue

            product_name = _get_cell(row, col_map.get("product_name"))
            if not product_name:
                continue

            line = {
                "product_name": str(product_name).strip(),
                "product_code": _get_cell(row, col_map.get("product_code")),
                "quantity": _to_number(_get_cell(row, col_map.get("quantity"))),
                "unit_price": _to_number(_get_cell(row, col_map.get("unit_price"))),
                "uom": _get_cell(row, col_map.get("uom")),
                "line_total": _to_number(_get_cell(row, col_map.get("line_total"))),
                "tax_rate": _to_number(_get_cell(row, col_map.get("tax_rate"))),
                "discount_rate": _to_number(_get_cell(row, col_map.get("discount_rate"))),
                "discount_amount": _to_number(_get_cell(row, col_map.get("discount_amount"))),
            }
            if not line["quantity"] or line["quantity"] <= 0:
                continue
            lines.append(line)

        if lines:
            orders.append({
                "sheet_name": sheet_name,
                "lines": lines,
                "total_amount": sum(l["line_total"] or 0 for l in lines),
            })

    wb.close()
    return orders


def _find_header_row(rows: list[tuple]) -> int | None:
    """Find the row index that looks like a header."""
    keywords = {"tên", "sản phẩm", "hàng hóa", "mặt hàng", "product", "item", "mã", "số lượng", "sl", "đơn giá", "stt"}
    for i, row in enumerate(rows[:20]):  # Only check first 20 rows
        if not row:
            continue
        text = " ".join(str(c).lower() for c in row if c)
        matches = sum(1 for kw in keywords if kw in text)
        if matches >= 2:
            return i
    return None


def _map_columns(headers: list[str]) -> dict[str, int | None]:
    """Map semantic column names to column indices."""
    mapping: dict[str, int | None] = {
        "product_name": None,
        "product_code": None,
        "quantity": None,
        "unit_price": None,
        "uom": None,
        "line_total": None,
        "tax_rate": None,
        "discount_rate": None,
        "discount_amount": None,
    }

    name_patterns = ["tên hàng", "tên sản phẩm", "tên mặt hàng", "hàng hóa", "tên hh", "diễn giải", "product name", "item"]
    code_patterns = ["mã hàng", "mã sp", "mã sản phẩm", "mã hh", "product code", "item code", "mã"]
    qty_patterns = ["số lượng", "sl", "qty", "quantity", "s.lượng"]
    price_patterns = ["đơn giá", "giá", "unit price", "price", "đ.giá"]
    uom_patterns = ["đvt", "đơn vị", "unit", "uom", "đơn vị tính"]
    total_patterns = ["thành tiền", "tổng", "amount", "total", "t.tiền"]
    tax_patterns = ["thuế", "vat", "tax", "%thuế"]
    discount_patterns = ["chiết khấu", "ck", "discount", "giảm giá"]

    for i, h in enumerate(headers):
        if not h:
            continue
        if mapping["product_name"] is None and any(p in h for p in name_patterns):
            mapping["product_name"] = i
        elif mapping["product_code"] is None and any(p in h for p in code_patterns):
            mapping["product_code"] = i
        elif mapping["quantity"] is None and any(p in h for p in qty_patterns):
            mapping["quantity"] = i
        elif mapping["unit_price"] is None and any(p in h for p in price_patterns):
            mapping["unit_price"] = i
        elif mapping["uom"] is None and any(p in h for p in uom_patterns):
            mapping["uom"] = i
        elif mapping["line_total"] is None and any(p in h for p in total_patterns):
            mapping["line_total"] = i
        elif mapping["tax_rate"] is None and any(p in h for p in tax_patterns):
            mapping["tax_rate"] = i
        elif mapping["discount_amount"] is None and any(p in h for p in discount_patterns):
            mapping["discount_amount"] = i

    return mapping


def _get_cell(row: tuple, idx: int | None) -> Any:
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _to_number(val: Any) -> float | None:
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        # Try stripping common formatting
        s = str(val).replace(",", "").replace(".", "").replace(" ", "").replace("đ", "")
        try:
            return float(s)
        except (ValueError, TypeError):
            return None


def import_excel_to_orders(db: Session, file_path: str, file_name: str) -> dict:
    """
    Parse Excel file and create ProcessedOrder + OrderLine records.
    Returns summary stats.
    """
    # Create a RawDocument record for tracking
    raw_doc = RawDocument(
        file_name=file_name,
        file_path=file_path,
        document_type="purchase_order",
        ocr_status="done",
    )
    db.add(raw_doc)
    db.flush()

    parsed_orders = parse_excel_orders(file_path)

    total_orders = 0
    total_products = 0
    total_unmapped = 0
    order_ids = []

    for po in parsed_orders:
        order = ProcessedOrder(
            raw_document_id=raw_doc.id,
            total_amount=po.get("total_amount"),
            status="processing",
            processed_at=datetime.utcnow(),
        )
        db.add(order)
        db.flush()

        for line_data in po["lines"]:
            temp_code, product_id = resolve_temp_code(
                db,
                line_data.get("product_code"),
                line_data["product_name"],
                line_data.get("tax_rate"),
            )
            mapping_status = "mapped" if product_id else "pending"
            if not product_id:
                total_unmapped += 1

            line = OrderLine(
                processed_order_id=order.id,
                temp_code=temp_code,
                product_id=product_id,
                product_name_original=line_data["product_name"],
                quantity=line_data.get("quantity"),
                unit_price=line_data.get("unit_price"),
                uom_original=line_data.get("uom"),
                line_total=line_data.get("line_total"),
                tax_rate=line_data.get("tax_rate"),
                discount_rate=line_data.get("discount_rate"),
                discount_amount=line_data.get("discount_amount"),
                mapping_status=mapping_status,
            )
            db.add(line)
            total_products += 1

        # Update order status based on mapping
        if total_unmapped == 0:
            order.status = "completed"
        order_ids.append(str(order.id))
        total_orders += 1

    db.commit()

    return {
        "file_name": file_name,
        "raw_document_id": str(raw_doc.id),
        "total_orders": total_orders,
        "total_products": total_products,
        "total_unmapped": total_unmapped,
        "order_ids": order_ids,
    }

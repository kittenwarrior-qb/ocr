import io
from datetime import datetime, timedelta
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy.orm import Session

from app.models.session import OcrSession
from app.models.document import RawDocument, ProcessedOrder


def _add_one_month(d):
    """Add one month to a date, clamping the day to the last valid day of the target month."""
    month = d.month + 1
    year = d.year
    if month > 12:
        month = 1
        year += 1
    # Clamp day to last day of target month
    if month == 12:
        next_month_first = d.replace(year=year + 1, month=1, day=1)
    else:
        next_month_first = d.replace(year=year, month=month + 1, day=1)
    last_day = (next_month_first - timedelta(days=1)).day
    day = min(d.day, last_day)
    return d.replace(year=year, month=month, day=day)


def create_session(db: Session, name: str, note: str | None = None) -> OcrSession:
    s = OcrSession(name=name, note=note)
    db.add(s)
    db.commit()
    db.refresh(s)
    return s


def list_sessions(db: Session) -> list[OcrSession]:
    return db.query(OcrSession).order_by(OcrSession.created_at.desc()).all()


def get_session(db: Session, session_id: UUID) -> OcrSession | None:
    return db.query(OcrSession).filter(OcrSession.id == session_id).first()


def close_session(db: Session, session_id: UUID) -> OcrSession | None:
    s = get_session(db, session_id)
    if s and s.status == "active":
        s.status = "closed"
        s.closed_at = datetime.utcnow()
        db.commit()
        db.refresh(s)
    return s


def reopen_session(db: Session, session_id: UUID) -> OcrSession | None:
    s = get_session(db, session_id)
    if s and s.status == "closed":
        s.status = "active"
        s.closed_at = None
        db.commit()
        db.refresh(s)
    return s


def session_stats(db: Session, session: OcrSession) -> dict:
    from app.models.document import ProcessedOrder
    docs = session.raw_documents
    doc_ids = [d.id for d in docs]
    order_count = 0
    if doc_ids:
        order_count = db.query(ProcessedOrder).filter(
            ProcessedOrder.raw_document_id.in_(doc_ids)
        ).count()
    return {
        "doc_count": len(docs),
        "done_count": sum(1 for d in docs if d.ocr_status == "done"),
        "order_count": order_count,
    }



def export_session_excel(db: Session, session_id: UUID) -> bytes:
    """Export session in MISA import format matching order_test.xlsx template.
    Sheet 1: 'Nhập khẩu Đơn hàng' (60 cols)
    Sheet 2: 'nhập khẩu hàng hóa' (16 cols)
    """
    from openpyxl.utils import get_column_letter

    session = get_session(db, session_id)
    if not session:
        raise ValueError("Session not found")

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Nhập khẩu Đơn hàng"
    ws2 = wb.create_sheet("nhập khẩu hàng hóa")

    # ── Sheet 1 headers (60 cols, matching order_test.xlsx) ──
    hdrs1 = [
        "Sử dụng ngoại tệ", "Loại tiền", "Tỷ giá",
        "Số đơn hàng (*)", "Ngày đặt hàng (*)", "Số PO", "Nhân viên bán hàng (*)",
        "Khách hàng", "Liên hệ", "Đơn hàng cha", "Báo giá", "Cơ hội", "Chiến dịch",
        "Giá trị đơn hàng (*)", "Giá trị thanh lý", "Diễn giải", "Loại đơn hàng (*)",
        "Số ngày được nợ", "Hạn giao hàng", "Hạn thanh toán (*)", "Tình trạng",
        "Ngày ghi sổ", "Thực thu", "Tình trạng giao hàng", "Dự kiến chi",
        "Tình trạng thanh toán", "Hạn sản xuất", "Đã xuất hóa đơn",
        "Khách hàng (Hóa đơn)", "Người mua hàng",
        "Quốc gia (Hóa đơn)", "Tỉnh/Thành phố (Hóa đơn)", "Quận/Huyện (Hóa đơn)",
        "Phường/Xã (Hóa đơn)", "Số nhà, Đường phố (Hóa đơn)", "Mã vùng (Hóa đơn)",
        "Địa chỉ (Hóa đơn)", "Người nhận hàng", "Điện thoại",
        "Quốc gia (Giao hàng)", "Tỉnh/Thành phố (Giao hàng)", "Quận/Huyện (Giao hàng)",
        "Phường/Xã (Giao hàng)", None, "Mã vùng (Giao hàng)",
        "Địa chỉ (Giao hàng)", "Mô tả", "Ghi chú hóa đơn", "Người thực hiện",
        "Dùng chung", "Ngừng theo dõi", "Đối tác/CTV giới thiệu",
        "Đồng bộ đơn giá sau CK", "Dự án bán hàng", "Nguồn gốc",
        "Nhân viên kho", "Nhân viên giao hàng", "Ngày giao dự kiến",
        "Tuyến vận chuyển", "Thực chi",
    ]
    widths1 = [
        16.82, 9.54, 7.0, 15.73, 17.54, 7.18, 22.73,
        12.27, 8.45, 13.82, 8.27, 7.45, 18.54,
        19.27, 15.0, 29.73, 17.54,
        16.45, 14.54, 18.54, 14.18,
        12.18, 9.73, 20.45, 12.18,
        21.27, 13.45, 16.27,
        22.18, 16.45,
        19.45, 25.82, 22.82,
        21.27, 28.73, 19.45,
        65.45, 17.0, 10.82,
        20.82, 27.27, 24.27,
        22.73, 30.18, 20.82,
        65.45, 36.27, 16.27, 23.18,
        12.45, 15.73, 22.18,
        24.0, 15.73, 11.27,
        14.82, 20.18, 18.18,
        18.18, 9.45,
    ]

    # ── Sheet 2 headers (16 cols, matching order_test.xlsx) ──
    hdrs2 = [
        "Mã hàng hóa", "Diễn giải", "Kho", "Đơn vị tính",
        "Số lượng", "Đơn giá", "Thành tiền", "Tỷ lệ chiết khấu",
        "Tiền chiết khấu", "Thuế suất", "Tiền thuế", "Tổng tiền",
        "Số lô", "Hạn sử dụng", "Hàng KM", "Đơn hàng (*)",
    ]
    widths2 = [
        13.27, 23.54, 6.45, 11.54,
        9.45, 8.45, 11.27, 16.27,
        15.82, 10.54, 10.27, 13.0,
        5.82, 13.18, 10.73, 13.27,
    ]

    # Style headers (bold Times New Roman, no fill)
    hfont = Font(name="Times New Roman", size=11, bold=True)
    for col, h in enumerate(hdrs1, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = hfont
    for col, w in enumerate(widths1, 1):
        ws1.column_dimensions[get_column_letter(col)].width = w

    for col, h in enumerate(hdrs2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = hfont
    for col, w in enumerate(widths2, 1):
        ws2.column_dimensions[get_column_letter(col)].width = w

    # ── Collect data ──
    order_row = 2
    item_row = 2

    for raw in session.raw_documents:
        if raw.ocr_status != "done":
            continue

        order = db.query(ProcessedOrder).filter(
            ProcessedOrder.raw_document_id == raw.id
        ).first()
        if not order:
            continue

        meta = order.extra_data or {}
        partner = order.partner
        partner_code = partner.code if partner else ""
        partner_name = partner.legal_name if partner else ""
        addr = order.delivery_address
        addr_text = addr.full_address if addr else ""

        # Use customer info from extra_data (set by frontend) if available
        customer_code = meta.get("customer_code") or partner_code
        customer_name = meta.get("customer_name") or partner_name

        order_date_str = order.order_date.strftime("%d/%m/%Y") if order.order_date else ""
        delivery_date_str = order.delivery_date.strftime("%d/%m/%Y") if order.delivery_date else ""
        # Hạn thanh toán (*) = ngày đặt hàng + 1 tháng
        if meta.get("payment_due"):
            payment_due_str = meta.get("payment_due")
        elif order.order_date:
            payment_due_str = _add_one_month(order.order_date).strftime("%d/%m/%Y")
        else:
            payment_due_str = ""

        # Sheet 1 row (60 cols)
        vals1 = [
            None,                                                       # 1  Sử dụng ngoại tệ
            None,                                                       # 2  Loại tiền
            None,                                                       # 3  Tỷ giá
            order.order_number or "",                                    # 4  Số đơn hàng
            order_date_str,                                             # 5  Ngày đặt hàng
            order.po_number or "",                                      # 6  Số PO
            meta.get("salesperson") or "KM-1989 Nguyễn Văn Ân",           # 7  Nhân viên bán hàng (*)
            customer_code,                                              # 8  Khách hàng
            meta.get("contact", ""),                                    # 9  Liên hệ
            meta.get("parent_order", ""),                               # 10 Đơn hàng cha
            meta.get("quotation", ""),                                  # 11 Báo giá
            meta.get("opportunity", ""),                                # 12 Cơ hội
            meta.get("campaign", ""),                                   # 13 Chiến dịch
            float(order.total_amount) if order.total_amount else None,  # 14 Giá trị đơn hàng
            meta.get("liquidation_value", ""),                          # 15 Giá trị thanh lý
            meta.get("description", ""),                                # 16 Diễn giải
            meta.get("order_type") or "Kênh MT",                        # 17 Loại đơn hàng (*)
            meta.get("credit_days", ""),                                # 18 Số ngày được nợ
            delivery_date_str,                                          # 19 Hạn giao hàng
            payment_due_str,                                            # 20 Hạn thanh toán (*)
            meta.get("status", "Chưa thực hiện"),                       # 21 Tình trạng
            meta.get("record_date", ""),                                # 22 Ngày ghi sổ
            meta.get("actual_revenue", ""),                             # 23 Thực thu
            meta.get("delivery_status", ""),                            # 24 Tình trạng giao hàng
            meta.get("expected_expense", ""),                           # 25 Dự kiến chi
            meta.get("payment_status", ""),                             # 26 Tình trạng thanh toán
            meta.get("production_deadline", ""),                        # 27 Hạn sản xuất
            meta.get("invoice_issued", ""),                             # 28 Đã xuất hóa đơn
            customer_code,                                              # 29 Khách hàng (Hóa đơn) - mã KH
            meta.get("invoice_buyer", ""),                              # 30 Người mua hàng
            meta.get("invoice_country", "Việt Nam"),                    # 31 Quốc gia (Hóa đơn)
            meta.get("invoice_city", ""),                               # 32 Tỉnh/Thành phố (Hóa đơn)
            meta.get("invoice_district", ""),                           # 33 Quận/Huyện (Hóa đơn)
            meta.get("invoice_ward", ""),                               # 34 Phường/Xã (Hóa đơn)
            meta.get("invoice_street", ""),                             # 35 Số nhà, Đường phố (Hóa đơn)
            meta.get("invoice_area_code", ""),                          # 36 Mã vùng (Hóa đơn)
            meta.get("invoice_address", ""),                            # 37 Địa chỉ (Hóa đơn)
            meta.get("delivery_receiver", ""),                          # 38 Người nhận hàng
            meta.get("delivery_phone", ""),                             # 39 Điện thoại
            meta.get("delivery_country", "Việt Nam"),                   # 40 Quốc gia (Giao hàng)
            meta.get("delivery_city", ""),                              # 41 Tỉnh/Thành phố (Giao hàng)
            meta.get("delivery_district", ""),                          # 42 Quận/Huyện (Giao hàng)
            meta.get("delivery_ward", ""),                              # 43 Phường/Xã (Giao hàng)
            None,                                                       # 44 (cột trống - MISA yêu cầu)
            meta.get("delivery_area_code", ""),                         # 45 Mã vùng (Giao hàng)
            meta.get("delivery_address", addr_text),                    # 46 Địa chỉ (Giao hàng)
            meta.get("note_description", ""),                           # 47 Mô tả
            meta.get("invoice_note", ""),                               # 48 Ghi chú hóa đơn
            meta.get("executor", ""),                                   # 49 Người thực hiện
            meta.get("shared", ""),                                     # 50 Dùng chung
            "",                                                         # 51 Ngừng theo dõi
            meta.get("referral_partner", ""),                           # 52 Đối tác/CTV giới thiệu
            meta.get("sync_price_after_discount", ""),                  # 53 Đồng bộ đơn giá sau CK
            meta.get("sales_project", ""),                              # 54 Dự án bán hàng
            meta.get("origin", ""),                                     # 55 Nguồn gốc
            meta.get("warehouse_staff", ""),                            # 56 Nhân viên kho
            meta.get("delivery_staff", ""),                             # 57 Nhân viên giao hàng
            meta.get("expected_delivery_date", ""),                     # 58 Ngày giao dự kiến
            meta.get("shipping_route", ""),                             # 59 Tuyến vận chuyển
            meta.get("actual_expense", ""),                             # 60 Thực chi
        ]
        for col, v in enumerate(vals1, 1):
            ws1.cell(row=order_row, column=col, value=v)
        order_row += 1

        # Sheet 2 rows (product lines)
        order_ref = order.order_number or ""
        for line in order.lines:
            product_code = ""
            product_name = line.product_name_original
            uom = line.uom_original or ""
            if line.product_id and line.product:
                product_code = line.product.code or ""
                product_name = line.product.display_name or line.product_name_original
                uom = line.product.uom or line.uom_original or ""
            else:
                # Fallback for unmapped lines: use OCR code so MISA has a reference
                product_code = line.ocr_product_code or line.temp_code or ""

            qty = float(line.quantity) if line.quantity is not None else None
            price = float(line.unit_price) if line.unit_price is not None else None
            amount = float(line.line_total) if line.line_total is not None else None
            if not price and amount and qty:
                price = round(amount / qty, 2)
            dk_rate = float(line.discount_rate) if getattr(line, 'discount_rate', None) is not None else 0
            dk_amt = float(line.discount_amount) if getattr(line, 'discount_amount', None) is not None else None
            if dk_amt is None:
                dk_amt = round((amount or 0) * dk_rate / 100, 2) if amount and dk_rate else 0
            tax_rate = float(line.tax_rate) if line.tax_rate is not None else 0
            base = (amount or 0) - (dk_amt or 0)
            tax_amt = round(base * tax_rate / 100, 2) if amount else 0
            total = (amount or 0) - (dk_amt or 0) + (tax_amt or 0) if amount else 0

            tax_str = f"{int(tax_rate)}%"

            ws2.cell(row=item_row, column=1, value=product_code)       # Mã hàng hóa
            ws2.cell(row=item_row, column=2, value=product_name)       # Diễn giải
            ws2.cell(row=item_row, column=3, value=None)               # Kho
            ws2.cell(row=item_row, column=4, value=uom)                # Đơn vị tính
            ws2.cell(row=item_row, column=5, value=qty)                # Số lượng
            ws2.cell(row=item_row, column=6, value=price)              # Đơn giá
            ws2.cell(row=item_row, column=7, value=amount)             # Thành tiền
            ws2.cell(row=item_row, column=8, value=dk_rate)            # Tỷ lệ chiết khấu
            ws2.cell(row=item_row, column=9, value=dk_amt)             # Tiền chiết khấu
            ws2.cell(row=item_row, column=10, value=tax_str)           # Thuế suất
            ws2.cell(row=item_row, column=11, value=tax_amt)           # Tiền thuế
            ws2.cell(row=item_row, column=12, value=total)             # Tổng tiền
            ws2.cell(row=item_row, column=13, value=None)              # Số lô
            ws2.cell(row=item_row, column=14, value=None)              # Hạn sử dụng
            ws2.cell(row=item_row, column=15, value=None)              # Hàng KM
            ws2.cell(row=item_row, column=16, value=order_ref)         # Đơn hàng (*)
            item_row += 1

    # Mark orders as exported
    for raw in session.raw_documents:
        if raw.ocr_status != "done":
            continue
        order = db.query(ProcessedOrder).filter(
            ProcessedOrder.raw_document_id == raw.id
        ).first()
        if order:
            order.status = "exported"
    db.commit()

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()



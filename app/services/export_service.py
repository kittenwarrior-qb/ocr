import io
from decimal import Decimal
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy.orm import Session

from app.models.document import BillLine, OrderLine, ProcessedBill, ProcessedOrder
from app.models.partner import Partner, PartnerAddress
from app.models.product import Product


def _get_product_name(db: Session, product_id) -> str:
    if not product_id:
        return ""
    p = db.query(Product).filter(Product.id == product_id).first()
    return p.display_name if p else ""


def _get_product_code(db: Session, product_id) -> str:
    if not product_id:
        return ""
    p = db.query(Product).filter(Product.id == product_id).first()
    return p.code if p else ""


def _get_product_uom(db: Session, product_id, fallback: str = "") -> str:
    if not product_id:
        return fallback or ""
    p = db.query(Product).filter(Product.id == product_id).first()
    return p.uom if p else fallback or ""


def _get_partner(db: Session, partner_id) -> Partner | None:
    if not partner_id:
        return None
    return db.query(Partner).filter(Partner.id == partner_id).first()


def _get_address(db: Session, addr_id) -> PartnerAddress | None:
    if not addr_id:
        return None
    return db.query(PartnerAddress).filter(PartnerAddress.id == addr_id).first()


def _style_header(ws, headers: list[str]) -> None:
    fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def _dec(val) -> float | None:
    if val is None:
        return None
    return float(val)


# ── Misa format ───────────────────────────────────────────────────────────────

MISA_ORDER_HEADERS = [
    "Ngày chứng từ", "Số chứng từ", "Mã đối tượng", "Tên đối tượng", "MST",
    "Địa chỉ giao", "Mã hàng", "Tên hàng", "ĐVT", "Số lượng",
    "Đơn giá", "Thành tiền",
]

MISA_BILL_HEADERS = [
    "Ngày hóa đơn", "Số hóa đơn", "Mã đối tượng", "Tên đối tượng", "MST",
    "Mã hàng", "Tên hàng", "ĐVT", "Số lượng", "Đơn giá",
    "Thành tiền", "Thuế suất (%)", "Tiền thuế",
]

BRAVO_ORDER_HEADERS = [
    "DocDate", "DocNo", "PartnerCode", "PartnerName", "TaxCode",
    "DeliveryAddress", "ItemCode", "ItemName", "Unit", "Qty",
    "UnitPrice", "Amount",
]

BRAVO_BILL_HEADERS = [
    "InvoiceDate", "InvoiceNo", "VendorCode", "VendorName", "TaxCode",
    "ItemCode", "ItemName", "Unit", "Qty", "UnitPrice",
    "Amount", "TaxRate", "TaxAmount",
]


def export_order_to_excel(db: Session, order_id: UUID, fmt: str = "misa") -> bytes:
    order = db.query(ProcessedOrder).filter(ProcessedOrder.id == order_id).first()
    if not order:
        raise ValueError(f"ProcessedOrder {order_id} not found")

    partner = _get_partner(db, order.partner_id)
    address = _get_address(db, order.delivery_address_id)

    # Use MISA template format (2 sheets)
    if fmt == "misa_template":
        return _export_order_misa_template(db, order, partner, address)

    wb = Workbook()
    ws = wb.active
    ws.title = "Đơn đặt hàng" if fmt == "misa" else "PurchaseOrder"

    headers = MISA_ORDER_HEADERS if fmt == "misa" else BRAVO_ORDER_HEADERS
    _style_header(ws, headers)

    doc_date = str(order.order_date) if order.order_date else ""
    partner_code = partner.code if partner else ""
    partner_name = partner.legal_name if partner else ""
    tax_code = partner.tax_code if partner else ""
    addr_text = address.full_address if address else ""

    for line in order.lines:
        row = [
            doc_date,
            order.order_number or "",
            partner_code,
            partner_name,
            tax_code,
            addr_text,
            _get_product_code(db, line.product_id) or line.temp_code,
            _get_product_name(db, line.product_id) or line.product_name_original,
            _get_product_uom(db, line.product_id, line.uom_original),
            _dec(line.quantity),
            _dec(line.unit_price),
            _dec(line.line_total),
        ]
        ws.append(row)

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    order.status = "exported"
    db.commit()
    return buf.getvalue()


def _export_order_misa_template(db, order, partner, address):
    """Export using MISA template: bold Times New Roman headers, exact col widths, all 61 cols."""
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    from app.utils.customer_matcher import find_best_customer

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Nhập khẩu Đơn hàng"

    hdrs1 = [
        "Số đơn hàng (*)", "Ngày đặt hàng (*)",
        "Số PO", "Nhân viên bán hàng (*)", "Khách hàng", "Liên hệ", "Đơn hàng cha",
        "Báo giá", "Cơ hội", "Chiến dịch", "Giá trị đơn hàng (*)", "Giá trị thanh lý",
        "Loại đơn hàng (*)", "Số ngày được nợ", "Hạn giao hàng", "Hạn thanh toán (*)",
        "Tình trạng (*)", "Ngày ghi sổ", "Thực thu", "Tình trạng giao hàng", "Dự kiến chi",
        "Hạn sản xuất", "Đã xuất hóa đơn", "Khách hàng (Hóa đơn)", "Người mua hàng",
        "Quốc gia (Hóa đơn)", "Tỉnh/Thành phố (Hóa đơn)", "Quận/Huyện (Hóa đơn)", "Phường/Xã (Hóa đơn)", "Số nhà, Đường phố (Hóa đơn)",
        "Mã vùng (Hóa đơn)", "Địa chỉ (Hóa đơn)", "Phân cụm hóa đơn", "Người nhận hàng", "Điện thoại",
        "Quốc gia (Giao hàng)", "Tỉnh/Thành phố (Giao hàng)", "Quận/Huyện (Giao hàng)", "Phường/Xã (Giao hàng)", "Số nhà, Đường phố (Giao hàng)",
        "Mã vùng (Giao hàng)", "Địa chỉ (Giao hàng)", "Mô tả", "Ghi chú Hóa đơn", "Người thực hiện",
        "Dùng chung", "Ngừng theo dõi", "Đối tác/CTV giới thiệu", "Đồng bộ đơn giá sau CK", "Dự án bán hàng",
        "Nguồn gốc", "Nhân viên kho", "Nhân viên giao hàng", "Ngày giao dự kiến", "Tuyến vận chuyển", "Thực chi",
    ]

    widths1 = [
        15.73, 17.54, 7.18, 22.73, 12.27, 8.45, 13.82, 8.27, 7.45, 18.54,
        19.27, 15.0, 17.54, 16.45, 14.54, 18.54, 14.18, 12.18, 9.73, 20.45,
        12.18, 13.45, 16.27, 22.18, 16.45, 19.45, 25.82, 22.82, 21.27, 28.73,
        19.45, 65.45, 18.18, 17.0, 10.82, 20.82, 27.27, 24.27, 22.73, 30.18,
        20.82, 65.45, 36.27, 16.82, 23.18, 12.45, 15.73, 22.18, 24.0, 15.73,
        11.27, 14.82, 20.18, 18.18, 18.18, 9.45,
    ]

    hfont = Font(name="Times New Roman", size=11, bold=True)
    for col, h in enumerate(hdrs1, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = hfont
        ws1.column_dimensions[get_column_letter(col)].width = widths1[col - 1]

    meta = order.extra_data or {} if order.extra_data else {}
    partner_code = partner.code if partner else ""
    partner_name = partner.legal_name if partner else ""
    addr_text = address.full_address if address else ""
    order_date_str = order.order_date.strftime("%d/%m/%Y") if order.order_date else ""
    delivery_date_str = order.delivery_date.strftime("%d/%m/%Y") if order.delivery_date else ""

    # Auto-map customer from catalog if not already set in meta
    if not meta.get("invoice_customer") and not partner_name:
        # Try to find customer from order description or recipient_name
        search_name = order.description or order.recipient_name or ""
        if search_name:
            matched = find_best_customer(search_name, addr_text)
            if matched:
                meta = dict(meta)  # copy to avoid mutating
                meta.setdefault("invoice_customer", matched.get("name", ""))
                meta.setdefault("invoice_buyer", matched.get("owner", ""))
                meta.setdefault("invoice_city", matched.get("invoice_city", ""))
                meta.setdefault("invoice_district", matched.get("invoice_district", ""))
                meta.setdefault("invoice_ward", matched.get("invoice_ward", ""))
                meta.setdefault("invoice_address", matched.get("invoice_address", ""))
                meta.setdefault("invoice_street", matched.get("invoice_address", ""))
                meta.setdefault("delivery_receiver", matched.get("owner", ""))
                meta.setdefault("delivery_phone", matched.get("phone", ""))
                meta.setdefault("delivery_city", matched.get("invoice_city", ""))
                meta.setdefault("delivery_district", matched.get("invoice_district", ""))
                meta.setdefault("delivery_ward", matched.get("invoice_ward", ""))
                meta.setdefault("delivery_street", matched.get("delivery_address", "") or matched.get("invoice_address", ""))
                meta.setdefault("delivery_address", matched.get("delivery_address", "") or matched.get("invoice_address", ""))
                meta.setdefault("executor", matched.get("owner", ""))
                if not partner_code:
                    partner_code = matched.get("code", "")
                if not partner_name:
                    partner_name = matched.get("name", "")
                # Store the real customer code from customers.json
                meta.setdefault("customer_code", matched.get("code", ""))

    # Use customer code from customers.json (via meta) if available, else partner.code from DB
    customer_code = meta.get("customer_code") or partner_code

    vals1 = [
        order.order_number or "",                                   # 1  So don hang
        order_date_str,                                             # 2  Ngay dat hang
        order.po_number or "",                                      # 3  So PO
        meta.get("salesperson", "KM1989-Nguyễn Văn Ân"),             # 4  Nhan vien ban hang (mac dinh)
        customer_code,                                              # 5  Khach hang (ma tu customers.json)
        meta.get("contact", ""),                                    # 6  Lien he
        meta.get("parent_order", ""),                               # 7  Don hang cha
        meta.get("quotation", ""),                                  # 8  Bao gia
        meta.get("opportunity", ""),                                # 9  Co hoi
        meta.get("campaign", ""),                                   # 10 Chien dich
        _dec(order.total_amount),                                   # 11 Gia tri don hang
        meta.get("liquidation_value", ""),                          # 12 Gia tri thanh ly
        meta.get("order_type", ""),                                 # 13 Loai don hang (Dien giai da bo)
        meta.get("credit_days", ""),                                # 14 So ngay duoc no
        delivery_date_str,                                          # 15 Han giao hang
        meta.get("payment_due", ""),                                # 16 Han thanh toan
        meta.get("status", "Chua thuc hien"),                       # 17 Tinh trang
        meta.get("record_date", ""),                                # 18 Ngay ghi so
        meta.get("actual_revenue", ""),                             # 19 Thuc thu
        meta.get("delivery_status", ""),                            # 20 Tinh trang giao hang
        meta.get("expected_expense", ""),                           # 21 Du kien chi
        meta.get("production_deadline", ""),                        # 22 Han san xuat (Tinh trang thanh toan da bo)
        meta.get("invoice_issued", ""),                             # 23 Da xuat hoa don
        meta.get("invoice_customer", partner_name),                 # 24 Khach hang (Hoa don) - tên KH cho template cũ
        meta.get("invoice_buyer", ""),                              # 25 Nguoi mua hang
        meta.get("invoice_country", "Viet Nam"),                    # 26 Quoc gia (Hoa don)
        meta.get("invoice_city", ""),                               # 27 Tinh/Thanh pho (Hoa don)
        meta.get("invoice_district", ""),                           # 28 Quan/Huyen (Hoa don)
        meta.get("invoice_ward", ""),                               # 29 Phuong/Xa (Hoa don)
        meta.get("invoice_street", ""),                             # 30 So nha Duong pho (Hoa don)
        meta.get("invoice_area_code", ""),                          # 31 Ma vung (Hoa don)
        meta.get("invoice_address", ""),                            # 32 Dia chi (Hoa don)
        meta.get("invoice_cluster", ""),                            # 33 Phan cum hoa don
        meta.get("delivery_receiver", ""),                          # 34 Nguoi nhan hang
        meta.get("delivery_phone", ""),                             # 35 Dien thoai
        meta.get("delivery_country", "Viet Nam"),                   # 36 Quoc gia (Giao hang)
        meta.get("delivery_city", ""),                              # 37 Tinh/Thanh pho (Giao hang)
        meta.get("delivery_district", ""),                          # 38 Quan/Huyen (Giao hang)
        meta.get("delivery_ward", ""),                              # 39 Phuong/Xa (Giao hang)
        meta.get("delivery_street", addr_text),                     # 40 So nha Duong pho (Giao hang)
        meta.get("delivery_area_code", ""),                         # 41 Ma vung (Giao hang)
        meta.get("delivery_address", addr_text),                    # 42 Dia chi (Giao hang)
        meta.get("note_description", ""),                           # 43 Mo ta
        meta.get("invoice_note", ""),                               # 44 Ghi chu Hoa don
        meta.get("executor", ""),                                   # 45 Nguoi thuc hien (dropdown NV)
        meta.get("shared", ""),                                     # 46 Dung chung
        "",                                                         # 47 Ngung theo doi
        meta.get("referral_partner", ""),                           # 48 Doi tac/CTV gioi thieu
        meta.get("sync_price_after_discount", ""),                  # 49 Dong bo don gia sau CK
        meta.get("sales_project", ""),                              # 50 Du an ban hang
        meta.get("origin", ""),                                     # 51 Nguon goc
        meta.get("warehouse_staff", ""),                            # 52 Nhan vien kho
        meta.get("delivery_staff", ""),                             # 53 Nhan vien giao hang
        meta.get("expected_delivery_date", ""),                     # 54 Ngay giao du kien
        meta.get("shipping_route", ""),                             # 55 Tuyen van chuyen
        meta.get("actual_expense", ""),                             # 56 Thuc chi
    ]
    for col, v in enumerate(vals1, 1):
        ws1.cell(row=2, column=col, value=v)

    # Sheet 2: hang hoa
    ws2 = wb.create_sheet("nhập khẩu hàng hóa")
    hdrs2 = [
        "Mã hàng hóa", "Số lượng", "Diễn giải", "Đơn vị tính",
        "Đơn giá", "Thành tiền", "Tỷ lệ chiết khấu", "Tiền chiết khấu",
        "Thuế suất", "Tiền thuế", "Tổng tiền", "Ghi chú", "Hàng KM", "Đơn hàng (*)",
    ]
    widths2 = [13.27, 9.45, 23.54, 11.54, 8.45, 11.27, 16.27, 15.82, 10.54, 10.27, 13.0, 8.45, 10.73, 13.27]
    for col, h in enumerate(hdrs2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = hfont
        ws2.column_dimensions[get_column_letter(col)].width = widths2[col - 1]

    order_ref = order.order_number or ""
    for idx, line in enumerate(order.lines, 2):
        product_code = _get_product_code(db, line.product_id) or line.ocr_product_code or line.temp_code
        product_name = _get_product_name(db, line.product_id) or line.product_name_original
        uom = _get_product_uom(db, line.product_id, line.uom_original)
        qty = _dec(line.quantity)
        price = _dec(line.unit_price)
        amount = _dec(line.line_total)
        dk_rate = _dec(line.discount_rate)
        dk_amt = _dec(line.discount_amount) or (round(amount * dk_rate / 100, 2) if amount and dk_rate else None)
        tax_rate = _dec(line.tax_rate)
        tax_amt = round((amount - (dk_amt or 0)) * tax_rate / 100, 2) if amount and tax_rate else None
        total = (amount or 0) - (dk_amt or 0) + (tax_amt or 0) if amount else None
        ws2.cell(row=idx, column=1, value=product_code)
        ws2.cell(row=idx, column=2, value=qty)
        ws2.cell(row=idx, column=3, value=product_name)
        ws2.cell(row=idx, column=4, value=uom)
        ws2.cell(row=idx, column=5, value=price)
        ws2.cell(row=idx, column=6, value=amount)
        ws2.cell(row=idx, column=7, value=dk_rate)
        ws2.cell(row=idx, column=8, value=dk_amt)
        ws2.cell(row=idx, column=9, value=f"{int(tax_rate)}%" if tax_rate else None)
        ws2.cell(row=idx, column=10, value=tax_amt)
        ws2.cell(row=idx, column=11, value=total)
        ws2.cell(row=idx, column=14, value=order_ref)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    order.status = "exported"
    db.commit()
    return buf.getvalue()


def export_orders_bulk(db: Session, order_ids: list[UUID], fmt: str = "misa_template") -> bytes:
    """Export multiple orders into a single Excel file (all orders in sheet1, all lines in sheet2)."""
    from openpyxl.utils import get_column_letter

    orders = db.query(ProcessedOrder).filter(ProcessedOrder.id.in_(order_ids)).all()
    if not orders:
        raise ValueError("No orders found")

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Nhập khẩu Đơn hàng"
    ws2 = wb.create_sheet("nhập khẩu hàng hóa")

    hdrs1 = [
        "Số đơn hàng (*)", "Ngày đặt hàng (*)",
        "Số PO", "Nhân viên bán hàng (*)", "Khách hàng", "Liên hệ", "Đơn hàng cha",
        "Báo giá", "Cơ hội", "Chiến dịch", "Giá trị đơn hàng (*)", "Giá trị thanh lý",
        "Loại đơn hàng (*)", "Số ngày được nợ", "Hạn giao hàng", "Hạn thanh toán (*)",
        "Tình trạng (*)", "Ngày ghi sổ", "Thực thu", "Tình trạng giao hàng", "Dự kiến chi",
        "Hạn sản xuất", "Đã xuất hóa đơn", "Khách hàng (Hóa đơn)", "Người mua hàng",
        "Quốc gia (Hóa đơn)", "Tỉnh/Thành phố (Hóa đơn)", "Quận/Huyện (Hóa đơn)", "Phường/Xã (Hóa đơn)", "Số nhà, Đường phố (Hóa đơn)",
        "Mã vùng (Hóa đơn)", "Địa chỉ (Hóa đơn)", "Phân cụm hóa đơn", "Người nhận hàng", "Điện thoại",
        "Quốc gia (Giao hàng)", "Tỉnh/Thành phố (Giao hàng)", "Quận/Huyện (Giao hàng)", "Phường/Xã (Giao hàng)", "Số nhà, Đường phố (Giao hàng)",
        "Mã vùng (Giao hàng)", "Địa chỉ (Giao hàng)", "Mô tả", "Ghi chú Hóa đơn", "Người thực hiện",
        "Dùng chung", "Ngừng theo dõi", "Đối tác/CTV giới thiệu", "Đồng bộ đơn giá sau CK", "Dự án bán hàng",
        "Nguồn gốc", "Nhân viên kho", "Nhân viên giao hàng", "Ngày giao dự kiến", "Tuyến vận chuyển", "Thực chi",
    ]
    widths1 = [
        15.73, 17.54, 7.18, 22.73, 12.27, 8.45, 13.82, 8.27, 7.45, 18.54,
        19.27, 15.0, 17.54, 16.45, 14.54, 18.54, 14.18, 12.18, 9.73, 20.45,
        12.18, 13.45, 16.27, 22.18, 16.45, 19.45, 25.82, 22.82, 21.27, 28.73,
        19.45, 65.45, 18.18, 17.0, 10.82, 20.82, 27.27, 24.27, 22.73, 30.18,
        20.82, 65.45, 36.27, 16.82, 23.18, 12.45, 15.73, 22.18, 24.0, 15.73,
        11.27, 14.82, 20.18, 18.18, 18.18, 9.45,
    ]
    hdrs2 = [
        "Mã hàng hóa", "Số lượng", "Diễn giải", "Đơn vị tính",
        "Đơn giá", "Thành tiền", "Tỷ lệ chiết khấu", "Tiền chiết khấu",
        "Thuế suất", "Tiền thuế", "Tổng tiền", "Ghi chú", "Hàng KM", "Đơn hàng (*)",
    ]
    widths2 = [13.27, 9.45, 23.54, 11.54, 8.45, 11.27, 16.27, 15.82, 10.54, 10.27, 13.0, 8.45, 10.73, 13.27]

    hfont = Font(name="Times New Roman", size=11, bold=True)
    for col, h in enumerate(hdrs1, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = hfont
        ws1.column_dimensions[get_column_letter(col)].width = widths1[col - 1]
    for col, h in enumerate(hdrs2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = hfont
        ws2.column_dimensions[get_column_letter(col)].width = widths2[col - 1]

    order_row = 2
    item_row = 2

    for order in orders:
        partner = _get_partner(db, order.partner_id)
        address = _get_address(db, order.delivery_address_id)
        meta = order.extra_data or {} if order.extra_data else {}
        partner_code = partner.code if partner else ""
        partner_name = partner.legal_name if partner else ""
        addr_text = address.full_address if address else ""
        order_date_str = order.order_date.strftime("%d/%m/%Y") if order.order_date else ""
        delivery_date_str = order.delivery_date.strftime("%d/%m/%Y") if order.delivery_date else ""

        customer_code = meta.get("customer_code") or partner_code

        vals = [
            order.order_number or "", order_date_str, order.po_number or "",
            meta.get("salesperson", ""), customer_code, meta.get("contact", ""),
            meta.get("parent_order", ""), meta.get("quotation", ""), meta.get("opportunity", ""),
            meta.get("campaign", ""), _dec(order.total_amount), meta.get("liquidation_value", ""),
            meta.get("order_type", ""),
            meta.get("credit_days", ""), delivery_date_str, meta.get("payment_due", ""),
            meta.get("status", "Chưa thực hiện"), meta.get("record_date", ""),
            meta.get("actual_revenue", ""), meta.get("delivery_status", ""),
            meta.get("expected_expense", ""),
            meta.get("production_deadline", ""), meta.get("invoice_issued", ""),
            meta.get("customer_code") or partner_code, meta.get("invoice_buyer", ""),
            meta.get("invoice_country", "Việt Nam"), meta.get("invoice_city", ""),
            meta.get("invoice_district", ""), meta.get("invoice_ward", ""),
            meta.get("invoice_street", ""), meta.get("invoice_area_code", ""),
            meta.get("invoice_address", ""), meta.get("invoice_cluster", ""),
            meta.get("delivery_receiver", ""), meta.get("delivery_phone", ""),
            meta.get("delivery_country", "Việt Nam"), meta.get("delivery_city", ""),
            meta.get("delivery_district", ""), meta.get("delivery_ward", ""),
            meta.get("delivery_street", addr_text), meta.get("delivery_area_code", ""),
            meta.get("delivery_address", addr_text), meta.get("note_description", ""),
            meta.get("invoice_note", ""), meta.get("executor", ""),
            meta.get("shared", ""), "", meta.get("referral_partner", ""),
            meta.get("sync_price_after_discount", ""), meta.get("sales_project", ""),
            meta.get("origin", ""), meta.get("warehouse_staff", ""),
            meta.get("delivery_staff", ""), meta.get("expected_delivery_date", ""),
            meta.get("shipping_route", ""), meta.get("actual_expense", ""),
        ]
        for col, v in enumerate(vals, 1):
            ws1.cell(row=order_row, column=col, value=v)
        order_row += 1

        # Lines
        for line in order.lines:
            product_code = _get_product_code(db, line.product_id) or line.ocr_product_code or line.temp_code
            product_name = _get_product_name(db, line.product_id) or line.product_name_original
            uom = _get_product_uom(db, line.product_id, line.uom_original)
            qty = _dec(line.quantity)
            price = _dec(line.unit_price)
            amount = _dec(line.line_total)
            dk_rate = _dec(line.discount_rate)
            dk_amt = _dec(line.discount_amount) or (round(amount * dk_rate / 100, 2) if amount and dk_rate else None)
            tax_rate = _dec(line.tax_rate)
            tax_amt = round((amount - (dk_amt or 0)) * tax_rate / 100, 2) if amount and tax_rate else None
            total = (amount or 0) - (dk_amt or 0) + (tax_amt or 0) if amount else None
            ws2.cell(row=item_row, column=1, value=product_code)
            ws2.cell(row=item_row, column=2, value=qty)
            ws2.cell(row=item_row, column=3, value=product_name)
            ws2.cell(row=item_row, column=4, value=uom)
            ws2.cell(row=item_row, column=5, value=price)
            ws2.cell(row=item_row, column=6, value=amount)
            ws2.cell(row=item_row, column=7, value=dk_rate)
            ws2.cell(row=item_row, column=8, value=dk_amt)
            ws2.cell(row=item_row, column=9, value=f"{int(tax_rate)}%" if tax_rate else None)
            ws2.cell(row=item_row, column=10, value=tax_amt)
            ws2.cell(row=item_row, column=11, value=total)
            ws2.cell(row=item_row, column=14, value=order.order_number or "")
            item_row += 1

        order.status = "exported"

    db.commit()
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def export_bill_to_excel(db: Session, bill_id: UUID, fmt: str = "misa") -> bytes:
    bill = db.query(ProcessedBill).filter(ProcessedBill.id == bill_id).first()
    if not bill:
        raise ValueError(f"ProcessedBill {bill_id} not found")

    partner = _get_partner(db, bill.partner_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Hóa đơn GTGT" if fmt == "misa" else "VendorBill"

    headers = MISA_BILL_HEADERS if fmt == "misa" else BRAVO_BILL_HEADERS
    _style_header(ws, headers)

    doc_date = str(bill.invoice_date) if bill.invoice_date else ""
    partner_code = partner.code if partner else ""
    partner_name = partner.legal_name if partner else ""
    tax_code = partner.tax_code if partner else ""

    for line in bill.lines:
        tax_rate = _dec(line.tax_rate)
        amount = _dec(line.line_total)
        tax_amount = round(amount * tax_rate / 100, 2) if amount and tax_rate else None
        row = [
            doc_date,
            bill.invoice_number or "",
            partner_code,
            partner_name,
            tax_code,
            _get_product_code(db, line.product_id) or line.temp_code,
            _get_product_name(db, line.product_id) or line.product_name_original,
            _get_product_uom(db, line.product_id, line.uom_original),
            _dec(line.quantity),
            _dec(line.unit_price),
            amount,
            tax_rate,
            tax_amount,
        ]
        ws.append(row)

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    bill.status = "exported"
    db.commit()
    return buf.getvalue()

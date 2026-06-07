"""
Parse Excel purchase order files into extracted-data structure matching OCR output.

Supports:
1. Satori internal template (ĐƠN ĐẶT HÀNG kênh GT) — fixed layout
2. KingKong/generic purchase order Excel — auto-detect
"""
import re
from pathlib import Path
import unicodedata


def _clean_num(val) -> float:
    if val is None:
        return 0.0
    s = str(val).replace(',', '').replace(' ', '').strip()
    try:
        return float(s)
    except (ValueError, TypeError):
        return 0.0


def _str(val) -> str:
    if val is None:
        return ''
    return str(val).strip()


def _norm(val) -> str:
    text = _str(val).lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = text.replace("đ", "d")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _parse_date(val) -> str | None:
    if not val:
        return None
    s = _str(val)
    m = re.search(r'(\d{1,2})[/\-](\d{1,2})[/\-](\d{4})', s)
    if m:
        d, mo, y = m.group(1), m.group(2), m.group(3)
        return f"{y}-{mo.zfill(2)}-{d.zfill(2)}"
    return None


def _is_satori_template(rows: list) -> bool:
    """Detect Satori internal order template by checking header structure."""
    normalized = _norm(' '.join(
        _str(c) for row in rows[:25] for c in row if c is not None
    ))
    return (
        'satori' in normalized and
        'don dat hang' in normalized and
        'ben mua' in normalized and
        'stt' in normalized and
        ('ten hang' in normalized or 'ten hang hoa' in normalized) and
        ('so luong' in normalized or 'luong hang dat' in normalized)
    )


# Cột bắt buộc phải xác định được vị trí — thiếu bất kỳ cột nào trong số này
# thì KHÔNG được tự ý đoán: phải raise lỗi để người dùng kiểm tra lại file,
# tuyệt đối không trích xuất "đại khái" rồi đưa ra số liệu sai.
_SATORI_REQUIRED_COLUMNS = (
    "stt", "product_name", "uom", "qty_ordered",
    "unit_price", "line_total_pretax", "tax_rate", "line_total_with_tax",
)

# Mỗi field khớp với nhãn cột thực tế trong dòng tiêu đề bằng cách so khớp
# (các) cụm từ đã chuẩn hoá — không phụ thuộc vào vị trí cột trong file.
# Thứ tự các cụm trong mỗi tuple = độ ưu tiên thử khớp.
_SATORI_COLUMN_LABEL_PATTERNS: dict[str, tuple[str, ...]] = {
    "stt": ("stt",),
    "product_name": ("ten hang hoa", "ten hang"),
    "spec": ("quy cach",),
    "uom": ("don vi tinh", "dvt"),
    "qty_ordered": ("so luong hang dat", "luong hang dat", "luong dat hang"),
    "qty_promo": ("so luong khuyen mai", "luong khuyen mai"),
    "qty_total": ("tong san luong", "tong so luong"),
    "unit_price": ("don gia ban", "don gia"),
    "line_total_pretax": ("thanh tien chua", "thanh tien chua thue", "thanh tien chua gtgt"),
    "tax_rate": ("muc thue suat", "thue suat"),
    "tax_amount": ("tien thue",),
    "line_total_with_tax": ("tong tien gom", "tong tien gom gtgt", "tong tien sau thue"),
}


def _map_satori_columns(header_row: tuple) -> dict[str, int]:
    """
    Xác định vị trí (index) thực tế của từng cột nghiệp vụ bằng cách so khớp
    nhãn đã chuẩn hoá trong dòng tiêu đề — KHÔNG dựa vào số thứ tự cột cố định.

    Nhờ vậy nếu gặp 1 file cùng mẫu Satori nhưng các cột bị xê dịch / chèn thêm /
    đổi thứ tự, parser vẫn đọc đúng cột vì tra theo tên, không theo vị trí.

    Nếu không khớp được đủ các cột bắt buộc (_SATORI_REQUIRED_COLUMNS), raise
    ValueError thay vì đọc đại — tránh việc âm thầm trả về dữ liệu sai lệch
    (đây chính là rủi ro "đúng định dạng nhưng khác vị trí cột" mà ta cần chặn).
    """
    cells_norm = [_norm(c) for c in header_row]
    col_map: dict[str, int] = {}
    used_indices: set[int] = set()

    for field, patterns in _SATORI_COLUMN_LABEL_PATTERNS.items():
        for pattern in patterns:
            found = None
            for ci, cell_norm in enumerate(cells_norm):
                if ci in used_indices:
                    continue
                if pattern in cell_norm:
                    found = ci
                    break
            if found is not None:
                col_map[field] = found
                used_indices.add(found)
                break

    missing = [f for f in _SATORI_REQUIRED_COLUMNS if f not in col_map]
    if missing:
        raise ValueError(
            "File Excel có vẻ đúng mẫu Satori (kênh GT) nhưng không xác định "
            f"được vị trí cột cho: {', '.join(missing)}. "
            "Để tránh trích xuất sai dữ liệu, hệ thống sẽ KHÔNG tự đoán vị trí cột — "
            "vui lòng kiểm tra lại tiêu đề cột trong file gốc (có thể tên cột đã đổi "
            "hoặc bị thiếu) trước khi nhập."
        )

    return col_map


def _parse_satori_template(ws) -> dict:
    """
    Parse Satori fixed-format order template.

    Vị trí cột nghiệp vụ KHÔNG được hardcode — được xác định động từ nhãn cột
    thực tế trong dòng tiêu đề (xem _map_satori_columns), nên nếu gặp file cùng
    mẫu nhưng các cột bị xê dịch vị trí, parser vẫn đọc đúng dữ liệu 100% thay
    vì âm thầm đọc nhầm cột.
    Header:
      R12: Bên mua / MST
      R13: Người liên hệ / Điện thoại
      R14: Địa chỉ giao dịch
      R15: Địa chỉ giao hàng
    Data rows: after header row containing 'STT' và 'Số lượng'
    """
    rows = list(ws.iter_rows(values_only=True))
    result = {
        "document_type": "purchase_order",
        "customer_name": None,
        "customer_tax_code": None,
        "order_number": None,
        "order_date": None,
        "recipient_name": None,
        "delivery_address": None,
        "items": [],
        "total_amount": None,
        "tax_amount": None,
    }

    # ── Header metadata ────────────────────────────────────────────────────────
    for row in rows[:20]:
        flat = ' '.join(_str(c) for c in row if c is not None)
        vals = [_str(c) for c in row]

        label = (vals[1] if len(vals) > 1 else '').lower()
        value = vals[3] if len(vals) > 3 else ''
        right_label = (vals[8] if len(vals) > 8 else '').lower()
        right_value = vals[9] if len(vals) > 9 else ''

        if 'bên mua' in label and value and not result['customer_name']:
            result['customer_name'] = value
        if 'mst' in right_label and right_value and not result['customer_tax_code']:
            m = re.search(r'\d{10,13}', right_value)
            result['customer_tax_code'] = m.group(0) if m else right_value
        if 'người liên hệ' in label and value and not result['recipient_name']:
            result['recipient_name'] = value
        if 'địa chỉ giao hàng' in label and value and not result['delivery_address']:
            result['delivery_address'] = value
        elif 'địa chỉ giao dịch' in label and value and not result['delivery_address']:
            result['delivery_address'] = value

        # Customer name: "Bên mua : <name>"
        m = re.search(r'Bên mua\s*:\s*(.+)', flat, re.I)
        if m and not result['customer_name']:
            name = m.group(1).strip()
            # Remove MST part if present
            name = re.sub(r'\s*MST\s*:.*', '', name).strip()
            if name:
                result['customer_name'] = name

        # Tax code: "MST : <code>"
        m = re.search(r'MST\s*:\s*(\d{10,13})', flat, re.I)
        if m and not result['customer_tax_code']:
            result['customer_tax_code'] = m.group(1).strip()

        # Contact: "Người liên hệ : <name>"
        m = re.search(r'Người liên hệ\s*:\s*([^Đ][^\t\n]+)', flat, re.I)
        if m and not result['recipient_name']:
            result['recipient_name'] = m.group(1).strip().split('\t')[0].strip()

        # Delivery address
        m = re.search(r'Địa chỉ giao hàng\s*:\s*(.+)', flat, re.I)
        if m and not result['delivery_address']:
            result['delivery_address'] = m.group(1).strip()
        elif not result['delivery_address']:
            m = re.search(r'Địa chỉ giao dịch\s*:\s*(.+)', flat, re.I)
            if m:
                result['delivery_address'] = m.group(1).strip()

    # ── Find header row: row containing 'STT' + 'Số lượng' + 'Tên hàng' ──────
    header_row_idx = None
    for idx, row in enumerate(rows):
        normalized = _norm(' '.join(_str(c) for c in row if c is not None))
        if 'stt' in normalized and ('so luong' in normalized or 'luong' in normalized) and ('ten hang' in normalized or 'ten hang hoa' in normalized):
            header_row_idx = idx
            break

    if header_row_idx is None:
        return result

    # Vị trí cột được tra theo NHÃN trong dòng tiêu đề — không hardcode index,
    # nên nếu file cùng mẫu nhưng cột bị xê dịch, vẫn đọc đúng cột. Nếu không
    # khớp đủ cột bắt buộc, _map_satori_columns raise lỗi rõ ràng thay vì đoán.
    col_map = _map_satori_columns(rows[header_row_idx])
    data_start = header_row_idx + 2  # skip label row + column-number row

    def col_val(vals: list, field: str):
        idx2 = col_map.get(field)
        return vals[idx2] if idx2 is not None and idx2 < len(vals) else None

    # ── Extract items with quantity > 0 ──────────────────────────────────────
    for row in rows[data_start:]:
        vals = list(row)
        if not any(v is not None for v in vals):
            continue

        # Check if this is a valid product row (cột STT = số thứ tự)
        stt_val = _str(col_val(vals, 'stt'))
        try:
            stt = int(float(stt_val))
        except (ValueError, TypeError):
            # Might be total row — capture total from "Tổng tiền gồm GTGT" column
            flat = ' '.join(_str(v) for v in vals if v is not None).lower()
            if 'tổng tiền' in flat:
                total = _clean_num(col_val(vals, 'line_total_with_tax'))
                if total > 0:
                    result['total_amount'] = total
            continue

        product_name = _str(col_val(vals, 'product_name'))
        if not product_name:
            continue

        uom = _str(col_val(vals, 'uom'))
        qty_ordered = _clean_num(col_val(vals, 'qty_ordered'))
        unit_price = _clean_num(col_val(vals, 'unit_price'))
        line_total_pretax = _clean_num(col_val(vals, 'line_total_pretax'))
        tax_rate_raw = _clean_num(col_val(vals, 'tax_rate'))

        # Only include rows where quantity was filled in (> 0)
        if qty_ordered <= 0:
            continue

        # Tax rate: stored as 0.08 → convert to 8
        tax_rate = tax_rate_raw * 100 if tax_rate_raw < 1 else tax_rate_raw

        result['items'].append({
            "product_name": product_name,
            "product_code": "",
            "quantity": qty_ordered,
            "uom": uom,
            "unit_price": unit_price if unit_price > 0 else None,
            "line_total": line_total_pretax if line_total_pretax > 0 else None,
            "tax_rate": tax_rate,
        })

    # Không tự tính total_amount nếu file không in sẵn — số tự tính có thể sai
    # (thiếu dòng khuyến mãi/chiết khấu không tính thuế, v.v.). Giữ null để user
    # biết cần kiểm tra lại file gốc thay vì tin vào một con số suy diễn.

    return result


def _parse_generic(ws) -> dict:
    """Generic parser for other Excel PO formats (KingKong, etc.)."""
    rows = list(ws.iter_rows(values_only=True))
    result = {
        "document_type": "purchase_order",
        "customer_name": None, "customer_tax_code": None,
        "order_number": None, "order_date": None,
        "delivery_address": None, "recipient_name": None,
        "items": [], "total_amount": None, "tax_amount": None,
    }

    # Metadata pass
    for row in rows:
        flat = ' '.join(_str(c) for c in row if c is not None)
        m = re.search(r'(?:Mã phiếu|Số đơn|PO|Order)[:\s]+([A-Z0-9\-_/]+)', flat, re.I)
        if m and not result['order_number']:
            result['order_number'] = m.group(1).strip()
        m = re.search(r'(?:Ngày|Date)[:\s]+(\d{1,2}[/\-]\d{1,2}[/\-]\d{4})', flat, re.I)
        if m and not result['order_date']:
            result['order_date'] = _parse_date(m.group(1))
        m = re.search(r'MST[:\s]+(\d{10,13})', flat, re.I)
        if m and not result['customer_tax_code']:
            result['customer_tax_code'] = m.group(1).strip()
        # Customer name
        for cell in row:
            s = _str(cell)
            if re.search(r'CÔNG TY|TNHH|MART|STORE|CO\.,', s, re.I) and len(s) > 5:
                if not result['customer_name']:
                    result['customer_name'] = s
                    break
        # Delivery address
        for i, cell in enumerate(row):
            if re.search(r'Địa chỉ giao hàng|Địa chỉ kho nhận|Ship to', _str(cell), re.I):
                for j in range(i+1, len(row)):
                    v = _str(row[j])
                    if v and len(v) > 10:
                        result['delivery_address'] = v
                        break

    # Find header row
    header_idx = None
    col_map: dict = {}
    for idx, row in enumerate(rows):
        flat_lower = ' '.join(_str(c).lower() for c in row if c is not None)
        if ('sản phẩm' in flat_lower or 'tên hàng' in flat_lower) and \
           ('số lượng' in flat_lower or 'qty' in flat_lower or 'sl' in flat_lower):
            header_idx = idx
            for ci, cell in enumerate(row):
                s = _str(cell).lower().strip()
                if re.search(r'mã hàng|product.?code|barcode', s):
                    col_map['product_code'] = ci
                elif re.search(r'tên|sản phẩm|product.?name', s) and 'product_name' not in col_map:
                    col_map['product_name'] = ci
                elif re.search(r'số lượng|sl|qty|quantity', s):
                    col_map['quantity'] = ci
                elif re.search(r'đvt|đơn vị|unit', s):
                    col_map['uom'] = ci
                elif re.search(r'đơn giá|unit.?price|price', s):
                    col_map['unit_price'] = ci
                elif re.search(r'thuế suất|vat.*%|tax.*rate|%', s):
                    col_map['tax_rate'] = ci
                elif re.search(r'thành tiền|total|amount', s) and 'line_total' not in col_map:
                    col_map['line_total'] = ci
            break

    if header_idx is not None:
        for row in rows[header_idx + 1:]:
            flat = ' '.join(_str(c) for c in row if c is not None)
            if re.search(r'tổng tiền|tổng cộng|cần trả', flat, re.I):
                for cell in row:
                    n = _clean_num(cell)
                    if n > 100000 and not result['total_amount']:
                        result['total_amount'] = n
                continue
            row_num = None
            for cell in row[:3]:
                try:
                    n = float(_str(cell))
                    if 0 < n <= 9999:
                        row_num = n
                        break
                except (ValueError, TypeError):
                    pass
            if row_num is None:
                continue

            def col_val(key):
                idx2 = col_map.get(key)
                return row[idx2] if idx2 is not None and idx2 < len(row) else None

            qty = _clean_num(col_val('quantity'))
            if qty <= 0:
                continue

            unit_price = _clean_num(col_val('unit_price'))
            line_total = _clean_num(col_val('line_total'))
            tax_rate_raw = _clean_num(col_val('tax_rate'))
            tax_rate = tax_rate_raw * 100 if tax_rate_raw < 1 else tax_rate_raw

            result['items'].append({
                "product_code": _str(col_val('product_code')),
                "product_name": _str(col_val('product_name')),
                "quantity": qty,
                "uom": _str(col_val('uom')),
                "unit_price": unit_price or None,
                "line_total": line_total or None,
                "tax_rate": tax_rate or None,
            })

    return result


def parse_excel_order(file_path: str) -> dict:
    """Parse a purchase order Excel file — auto-detect format."""
    import openpyxl
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 25), values_only=True))
    if _is_satori_template(rows):
        return _parse_satori_template(ws)
    else:
        return _parse_generic(ws)
